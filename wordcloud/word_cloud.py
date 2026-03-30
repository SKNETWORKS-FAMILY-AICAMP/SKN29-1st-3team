from openpyxl import load_workbook
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from collections import Counter
import mecab_ko as MeCab

# =========================
# 형태소 분석기 초기화
# =========================
# 프로그램 시작 시 1회 생성 (성능 중요)
tagger = MeCab.Tagger()

# =========================
# 엑셀에서 텍스트 로드 + 형태소 분석
# =========================
def load_text_from_excel(filename="result.xlsx"):
    wb = load_workbook(filename)
    ws = wb.active

    tokens = []
    
    # 2행부터 순회 (1행은 헤더)
    for row in ws.iter_rows(min_row=2, values_only=True):
        content = row[2]  # content 컬럼 (3번째 컬럼)

        if content:
            # 각 문서별 키워드 추출 후 리스트에 누적
            # extend → 리스트 평탄화 (중요)
            tokens.extend(extract_keywords(content))

    return tokens  # 최종: 단어 리스트

# =========================
# 워드클라우드 생성
# =========================
def generate_wordcloud(words):
    # 단어 빈도 계산
    word_count = Counter(words)

    # 워드클라우드 객체 생성
    wc = WordCloud(
        font_path="/System/Library/Fonts/AppleSDGothicNeo.ttc",  # 한글 폰트 (없으면 깨짐)
        width=1000,
        height=500,
        background_color="white"
    ).generate_from_frequencies(word_count)

    # 화면 출력
    plt.figure(figsize=(12, 6))
    plt.imshow(wc)
    plt.axis("off")
    plt.show()

    # 이미지 파일 저장
    wc.to_file("wordcloud.png")
    print("wordcloud.png 생성 완료")


# =========================
# Mecab 형태소 분석 기반 키워드 추출
# =========================
def extract_keywords(text):
    tokens = []

    # Mecab 결과를 라인 단위로 파싱
    for line in tagger.parse(text).split("\n"):
        if line == "EOS":
            continue
        
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        
        word = parts[0]  # 실제 단어
        pos = parts[1].split(",")[0]  # 품사 (NNG, NNP 등)

        # 명사만 필터링
        # NNG: 일반명사 / NNP: 고유명사
        if pos in ["NNG", "NNP"]:  # ⚠ ".NNP" → 오타였음
            tokens.append(word)

    return tokens


# =========================
# 실행 진입점
# =========================
if __name__ == "__main__":
    # 엑셀 → 형태소 분석 → 토큰 리스트 생성
    words = load_text_from_excel("result.xlsx")

    # 워드클라우드 생성
    generate_wordcloud(words)